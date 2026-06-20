from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .document_models import CellResult, DocumentResult, ReviewStatus, TableResult


_LOW_CONFIDENCE_FILL = PatternFill("solid", fgColor="FFF2CC")
_REVIEW_REQUIRED_FILL = PatternFill("solid", fgColor="F4CCCC")
_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


class TableWorkbookExporter:
    """Render structured table results to a deterministic, editable workbook."""

    def export(self, document: DocumentResult, destination: str | Path) -> Path:
        if not document.tables:
            raise ValueError("document does not contain tables")

        output_path = Path(destination).resolve()
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)
        for index, table in enumerate(document.tables, start=1):
            sheet = workbook.create_sheet(self._sheet_title(table, index))
            self._render_table(sheet, table)

        workbook.properties.title = document.source_name
        workbook.properties.subject = "Customs Agent structured table export"
        workbook.save(output_path)
        return output_path

    def _render_table(self, sheet, table: TableResult) -> None:
        for cell_result in sorted(
            table.cells, key=lambda item: (item.row, item.column)
        ):
            row = cell_result.row + 1
            column = cell_result.column + 1
            cell = sheet.cell(
                row=row,
                column=column,
                value=self._safe_value(cell_result),
            )
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )
            self._apply_review_style(cell, cell_result)

            if cell_result.row_span > 1 or cell_result.column_span > 1:
                sheet.merge_cells(
                    start_row=row,
                    start_column=column,
                    end_row=row + cell_result.row_span - 1,
                    end_column=column + cell_result.column_span - 1,
                )

        for column in range(1, table.columns + 1):
            values = [
                str(sheet.cell(row=row, column=column).value or "")
                for row in range(1, table.rows + 1)
            ]
            width = min(max((len(value) for value in values), default=8) + 2, 50)
            sheet.column_dimensions[get_column_letter(column)].width = max(width, 10)

        for row in range(1, table.rows + 1):
            sheet.row_dimensions[row].height = 24
        sheet.freeze_panes = "A2" if table.rows > 1 else None
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(table.columns)}{table.rows}"
        )

    @staticmethod
    def _safe_value(cell: CellResult):
        value = (
            cell.normalized_value
            if cell.normalized_value is not None
            else cell.text
        )
        if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
            return f"'{value}"
        return value

    @staticmethod
    def _apply_review_style(cell, result: CellResult) -> None:
        if result.review_status == ReviewStatus.REVIEW_REQUIRED:
            cell.fill = _REVIEW_REQUIRED_FILL
        elif result.confidence is not None and result.confidence < 0.8:
            cell.fill = _LOW_CONFIDENCE_FILL
        elif result.row == 0:
            cell.fill = _HEADER_FILL
            cell.font = Font(bold=True)

    @staticmethod
    def _sheet_title(table: TableResult, index: int) -> str:
        raw_title = str(table.metadata.get("name") or f"Table {index}")
        for character in '[]:*?/\\':
            raw_title = raw_title.replace(character, "_")
        return raw_title[:31] or f"Table {index}"
